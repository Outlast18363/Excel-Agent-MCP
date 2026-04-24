"""Excel-facing service layer for the Excel MCP server."""

from __future__ import annotations

import contextlib
import io
import tempfile
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from PIL import Image

from .helpers import (
    ExcelServiceError,
    apply_number_format,
    apply_style,
    build_style_lookup,
    build_trace_node_payload,
    column_number_to_name,
    default_screenshot_output_path,
    expand_formulas_ref,
    extract_excel_error,
    format_formulas_ref,
    get_address,
    get_range_hidden_flags,
    get_merged_ranges,
    get_range_geometry,
    normalize_matrix_input,
    normalize_formula_query,
    normalize_formula_grid,
    normalize_number_format_value,
    normalize_number_format_grid,
    normalize_range_read_matrix,
    normalize_trace_ref,
    read_number_format,
    row_column_to_a1_address,
    safe_count,
    sheet_visible,
)
from .types import JsonValue


@dataclass(slots=True)
class WorkbookSession:
    """Track a live workbook registered with the MCP process."""

    workbook_id: str
    workbook: Any
    app: Any
    path: str
    read_only: bool
    visible: bool


class ExcelService:
    """Own workbook registry and live Excel interactions for MCP tools."""

    def __init__(self) -> None:
        """Initialize the service registry and app cache."""
        self._workbooks: dict[str, WorkbookSession] = {}
        self._path_index: dict[str, str] = {}
        self._apps: dict[bool, Any] = {}
        self._next_workbook_number = 1

    def open_workbook(
        self,
        *,
        path: str,
        read_only: bool = False,
        visible: bool = False,
        create_if_missing: bool = False,
    ) -> dict[str, JsonValue]:
        xw = self._require_xlwings()
        workbook_path = Path(path).expanduser()
        resolved_path = str(workbook_path.resolve(strict=False))

        existing_id = self._path_index.get(resolved_path)
        if existing_id:
            existing_session = self._workbooks.get(existing_id)
            if existing_session is not None:
                return self._session_payload(existing_session)

        app = self._get_or_create_app(visible=visible)

        if workbook_path.exists():
            workbook = self._find_open_workbook(app=app, target_path=resolved_path)
            if workbook is None:
                workbook = app.books.open(resolved_path, read_only=read_only)
        else:
            if not create_if_missing:
                raise ExcelServiceError(f"Workbook does not exist: {resolved_path}")

            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = app.books.add()
            workbook.save(resolved_path)

        workbook_id = self._new_workbook_id()
        session = WorkbookSession(
            workbook_id=workbook_id,
            workbook=workbook,
            app=app,
            path=resolved_path,
            read_only=bool(read_only),
            visible=bool(visible),
        )
        self._workbooks[workbook_id] = session
        self._path_index[resolved_path] = workbook_id
        return self._session_payload(session)

    def get_sheet_state(
        self,
        *,
        workbook_id: str,
        sheet: str,
        include_used_range: bool = True,
        include_hidden: bool = True,
        include_merged_ranges: bool = True,
        include_formula_stats: bool = True,
        include_object_counts: bool = True,
    ) -> dict[str, JsonValue]:
        session = self._get_workbook_session(workbook_id)
        worksheet = self._get_sheet(workbook_id=workbook_id, sheet_name=sheet)
        snapshot_path = self._create_sheet_state_snapshot(session)
        workbook_formula = None

        try:
            # Read structural metadata from an on-disk workbook snapshot so the
            # result does not depend on slow per-cell COM round-trips.
            workbook_formula = load_workbook(
                snapshot_path,
                read_only=False,
                data_only=False,
                keep_links=False,
            )
            worksheet_formula = workbook_formula[sheet]

            used_range = worksheet_formula.calculate_dimension()
            start_col, start_row, max_col, max_row = range_boundaries(used_range)

            data: dict[str, JsonValue] = {
                "sheet": worksheet_formula.title,
                "visible": worksheet_formula.sheet_state == "visible",
            }

            if include_used_range:
                data["used_range"] = used_range
                data["max_row"] = max_row
                data["max_col"] = max_col

            if include_hidden:
                data["hidden_rows"] = [
                    row_number
                    for row_number in range(start_row, max_row + 1)
                    if bool(worksheet_formula.row_dimensions[row_number].hidden)
                ]
                data["hidden_columns"] = [
                    column_number_to_name(column_number)
                    for column_number in range(start_col, max_col + 1)
                    if bool(
                        worksheet_formula.column_dimensions[
                            column_number_to_name(column_number)
                        ].hidden
                    )
                ]

            if include_merged_ranges:
                merged_ranges = sorted(str(cell_range) for cell_range in worksheet_formula.merged_cells.ranges)
                data["merged_ranges"] = merged_ranges
                data["merged_range_count"] = len(merged_ranges)

            if include_formula_stats:
                formula_count = 0
                nonempty_count = 0
                for formula_row in worksheet_formula.iter_rows(
                    min_row=start_row,
                    max_row=max_row,
                    min_col=start_col,
                    max_col=max_col,
                ):
                    for formula_cell in formula_row:
                        if formula_cell.data_type == "f":
                            formula_count += 1
                        if formula_cell.value not in (None, ""):
                            nonempty_count += 1
                data["formula_count"] = formula_count
                data["nonempty_cell_count"] = nonempty_count

            if include_object_counts:
                # Keep Excel-native object counts as an optional live read because
                # openpyxl does not provide a reliable sheet-level shape inventory.
                data["chart_count"] = safe_count(lambda: int(worksheet.api.ChartObjects().Count))
                data["shape_count"] = safe_count(lambda: int(worksheet.api.Shapes.Count))

            return data
        finally:
            if workbook_formula is not None:
                workbook_formula.close()
            if snapshot_path != Path(session.path):
                snapshot_path.unlink(missing_ok=True)

    def search_cell(
        self,
        *,
        workbook_id: str,
        query: str | int | float,
        sheet: str | None = None,
        limit: int = 10,
        match_formulas: bool = True,
    ) -> dict[str, JsonValue]:
        session = self._get_workbook_session(workbook_id)
        normalized_limit = self._normalize_search_limit(limit)
        scope = "sheet" if sheet is not None else "workbook"
        target_sheets = [self._get_sheet(workbook_id=workbook_id, sheet_name=sheet).name] if sheet else None
        kind, prepared_query = self._prepare_search_query(query)
        snapshot_path = self._create_sheet_state_snapshot(session)
        workbook_formula = None
        workbook_values = None

        try:
            workbook_formula = load_workbook(
                snapshot_path,
                read_only=False,
                data_only=False,
                keep_links=False,
            )
            workbook_values = load_workbook(
                snapshot_path,
                read_only=False,
                data_only=True,
                keep_links=False,
            )

            matches: list[str] = []
            truncated = False
            for sheet_name in target_sheets or workbook_formula.sheetnames:
                worksheet_formula = workbook_formula[sheet_name]
                worksheet_values = workbook_values[sheet_name]
                used_range = worksheet_formula.calculate_dimension()
                start_col, start_row, max_col, max_row = range_boundaries(used_range)

                for row_number, (formula_row, value_row) in enumerate(
                    zip(
                        worksheet_formula.iter_rows(
                            min_row=start_row,
                            max_row=max_row,
                            min_col=start_col,
                            max_col=max_col,
                        ),
                        worksheet_values.iter_rows(
                            min_row=start_row,
                            max_row=max_row,
                            min_col=start_col,
                            max_col=max_col,
                        ),
                    ),
                    start=start_row,
                ):
                    for col_number, (formula_cell, value_cell) in enumerate(
                        zip(formula_row, value_row),
                        start=start_col,
                    ):
                        if not self._search_cell_matches(
                            kind=kind,
                            prepared_query=prepared_query,
                            value=value_cell.value,
                            formula=formula_cell.value if formula_cell.data_type == "f" else None,
                            match_formulas=match_formulas,
                        ):
                            continue

                        if len(matches) >= normalized_limit:
                            truncated = True
                            break

                        address = row_column_to_a1_address(row_number, col_number)
                        matches.append(
                            address if scope == "sheet" else f"{worksheet_formula.title}!{address}"
                        )
                    if truncated:
                        break
                if truncated:
                    break

            data: dict[str, JsonValue] = {
                "query": query,
                "kind": kind,
                "scope": scope,
                "limit": normalized_limit,
                "count": len(matches),
                "truncated": truncated,
                "matches": matches,
            }
            if scope == "sheet":
                data["sheet"] = target_sheets[0]
            return data
        finally:
            if workbook_formula is not None:
                workbook_formula.close()
            if workbook_values is not None:
                workbook_values.close()
            if snapshot_path != Path(session.path):
                snapshot_path.unlink(missing_ok=True)

    def get_range(
        self,
        *,
        workbook_id: str,
        sheet: str,
        range_address: str,
        include_values: bool = True,
        include_formulas: bool = False,
        include_number_formats: bool = False,
        include_styles: bool = False,
        include_geometry: bool = False,
        include_hidden_flags: bool = False,
        include_merged_info: bool = False,
    ) -> dict[str, JsonValue]:
        target_range = self._get_range(workbook_id=workbook_id, sheet_name=sheet, range_address=range_address)
        rows = int(target_range.rows.count)
        cols = int(target_range.columns.count)

        data: dict[str, JsonValue] = {
            "sheet": sheet,
            "range": get_address(target_range),
            "rows": rows,
            "columns": cols,
        }

        if include_values:
            values_matrix = normalize_range_read_matrix(
                target_range.options(ndim=2, chunksize=10_000).value,
                rows,
                cols,
                "values",
            )
            data["values"] = values_matrix

        if include_formulas:
            data["formulas"] = normalize_formula_grid(target_range.formula, rows, cols)

        if include_number_formats:
            data["number_formats"] = self._get_number_format_matrix(
                target_range=target_range,
                rows=rows,
                cols=cols,
            )

        if include_styles:
            style_table, style_ids = build_style_lookup(target_range, rows, cols)
            data["style_table"] = style_table
            data["style_ids"] = style_ids

        if include_hidden_flags:
            row_hidden, column_hidden = get_range_hidden_flags(target_range, rows, cols)
            data["row_hidden"] = row_hidden
            data["column_hidden"] = column_hidden

        if include_merged_info:
            data["merged_ranges"] = get_merged_ranges(target_range)

        if include_geometry:
            data["geometry"] = get_range_geometry(target_range)

        return data

    def set_range(
        self,
        *,
        workbook_id: str,
        sheet: str,
        range_address: str,
        values: Any = None,
        formulas: Any = None,
        number_format: Any = None,
        style: dict[str, Any] | None = None,
        clear_contents: bool = False,
        clear_formats: bool = False,
        save_after: bool = False,
    ) -> dict[str, JsonValue]:
        session = self._get_workbook_session(workbook_id)
        target_range = self._get_range(workbook_id=workbook_id, sheet_name=sheet, range_address=range_address)
        rows = int(target_range.rows.count)
        cols = int(target_range.columns.count)

        values_matrix = (
            normalize_matrix_input(values, rows, cols, "values")
            if values is not None
            else None
        )
        formulas_matrix = (
            normalize_matrix_input(formulas, rows, cols, "formulas")
            if formulas is not None
            else None
        )

        if clear_contents:
            target_range.clear_contents()

        if clear_formats:
            target_range.clear_formats()

        updated_values = False
        if values_matrix is not None:
            target_range.value = values_matrix
            updated_values = True

        updated_formulas = False
        if formulas_matrix is not None:
            target_range.formula = formulas_matrix
            updated_formulas = True

        updated_style = False
        if number_format is not None:
            apply_number_format(target_range, number_format, rows, cols)
            updated_style = True

        if style:
            apply_style(target_range, style)
            updated_style = True

        saved = False
        if save_after:
            session.workbook.save()
            saved = True

        return {
            "sheet": sheet,
            "range": get_address(target_range),
            "updated_values": updated_values,
            "updated_formulas": updated_formulas,
            "updated_style": updated_style,
            "saved": saved,
        }

    def recalculate(
        self,
        *,
        workbook_id: str,
        scope: str = "workbook",
        sheet: str | None = None,
        range_address: str | None = None,
        scan_errors: bool = True,
        return_formula_stats: bool = True,
        max_error_locations_per_type: int = 50,
    ) -> dict[str, JsonValue]:
        session = self._get_workbook_session(workbook_id)
        session.app.calculate()
        targets = self._recalc_targets(workbook_id=workbook_id, scope=scope, sheet=sheet, range_address=range_address)

        total_formulas = 0
        error_summary: dict[str, dict[str, JsonValue]] = {}
        total_errors = 0

        if scan_errors or return_formula_stats:
            for target in targets:
                for cell in target:
                    formula = cell.formula
                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue

                    total_formulas += 1
                    if not scan_errors:
                        continue

                    error_literal = extract_excel_error(cell.value)
                    if error_literal is None:
                        continue

                    total_errors += 1
                    bucket = error_summary.setdefault(
                        error_literal,
                        {"count": 0, "locations": []},
                    )
                    bucket["count"] = int(bucket["count"]) + 1
                    locations = bucket["locations"]
                    if (
                        isinstance(locations, list)
                        and len(locations) < max_error_locations_per_type
                    ):
                        locations.append(get_address(cell))

        data: dict[str, JsonValue] = {
            "scope": scope,
            "recalculated": True,
        }

        if sheet is not None:
            data["sheet"] = sheet
        if range_address is not None:
            data["range"] = range_address
        if return_formula_stats:
            data["total_formulas"] = total_formulas
        if scan_errors:
            data["total_errors"] = total_errors
            data["error_summary"] = error_summary

        return data

    def local_screenshot(
        self,
        *,
        workbook_id: str,
        sheet: str,
        range_address: str,
        output_path: str | None = None,
    ) -> dict[str, JsonValue]:
        target_range = self._get_range(workbook_id=workbook_id, sheet_name=sheet, range_address=range_address)

        captured_range = get_address(target_range)
        # Default to a stable repo-local output path so MCP callers can reuse the
        # returned image path without relying on ephemeral temp files.
        target_path = (
            Path(output_path).expanduser()
            if output_path
            else default_screenshot_output_path(
                workbook_id=workbook_id,
                sheet=sheet,
                range_address=captured_range,
            )
        )
        target_path.parent.mkdir(parents=True, exist_ok=True)

        # Use xlwings native to_png method
        target_range.to_png(str(target_path))

        # xlwings produces RGBA PNGs where unfilled cells are fully transparent,
        # making text invisible on non-white viewers. Composite onto a white
        # background and convert to RGB so the result is fully opaque.
        img = Image.open(str(target_path))
        if img.mode == "RGBA":
            background = Image.new("RGBA", img.size, (255, 255, 255, 255))
            composited = Image.alpha_composite(background, img)
            composited.convert("RGB").save(str(target_path))

        data: dict[str, JsonValue] = {
            "sheet": sheet,
            "range": captured_range,
            "image_path": str(target_path),
        }

        return data

    def close_workbook(
        self,
        *,
        workbook_id: str,
        save: bool = False,
    ) -> dict[str, JsonValue]:
        """Close one managed workbook and release its Excel app when unused.

        Parameters:
            workbook_id: The workbook handle returned by ``open_workbook``.
            save: Whether to save pending edits before closing the workbook.

        Returns:
            A JSON-safe payload describing whether the workbook was saved,
            closed, and whether its backing Excel app was also shut down.
        """

        session = self._get_workbook_session(workbook_id)
        if save and session.read_only:
            raise ExcelServiceError("Cannot save a read-only workbook while closing it.")

        if save:
            try:
                session.workbook.save()
            except Exception as exc:
                raise ExcelServiceError(
                    f"Workbook `{workbook_id}` could not be saved before closing."
                ) from exc

        try:
            session.workbook.close()
        except Exception as exc:
            raise ExcelServiceError(f"Workbook `{workbook_id}` could not be closed.") from exc

        self._workbooks.pop(workbook_id, None)
        if self._path_index.get(session.path) == workbook_id:
            self._path_index.pop(session.path, None)

        app_closed = False
        has_other_sessions = any(other.app is session.app for other in self._workbooks.values())
        if not has_other_sessions:
            # Drop the app from the cache so later opens create a fresh handle.
            self._apps.pop(session.visible, None)
            try:
                session.app.quit()
                app_closed = True
            except Exception:
                app_closed = False

        return {
            "workbook_id": workbook_id,
            "path": session.path,
            "saved": bool(save),
            "closed": True,
            "app_closed": app_closed,
        }

    def trace_formula(
        self,
        *,
        workbook_id: str,
        sheet: str,
        range_address: str,
        direction: str,
        max_depth: int | None = 1,
        include_addresses: bool = True,
    ) -> dict[str, JsonValue]:
        """Trace formula precedents or dependents for a cell or range.

        Parameters:
            workbook_id: The workbook handle returned by ``open_workbook``.
            sheet: The sheet name containing the target range.
            range_address: The A1 target cell or range to trace.
            direction: Either ``precedents`` or ``dependents``.
            max_depth: The maximum traversal depth, or ``None`` for full expansion.
            include_addresses: Whether to include split sheet and range metadata.

        Returns:
            A JSON-safe trace payload containing the native dependency graph slice.
        """

        session = self._get_workbook_session(workbook_id)
        target_range = self._get_range(
            workbook_id=workbook_id,
            sheet_name=sheet,
            range_address=range_address,
        )
        normalized_direction = self._normalize_trace_direction(direction)
        normalized_depth = self._normalize_trace_depth(max_depth)
        trace_model, workbook_name = self._build_trace_model(session)
        sheet_name_map = {
            worksheet.name.upper(): worksheet.name
            for worksheet in session.workbook.sheets
        }
        root_refs = self._build_trace_root_refs(
            workbook_name=workbook_name,
            sheet=sheet,
            target_range=target_range,
        )
        trace_nodes, trace_edges = self._collect_trace_graph(
            trace_model=trace_model,
            root_refs=root_refs,
            direction=normalized_direction,
            max_depth=normalized_depth,
            active_sheet=sheet,
            sheet_name_map=sheet_name_map,
            include_addresses=include_addresses,
        )

        return {
            "sheet": sheet,
            "range": get_address(target_range),
            "direction": normalized_direction,
            "max_depth": normalized_depth,
            "complete": True,
            "nodes": trace_nodes,
            "edges": trace_edges,
        }

    def _require_xlwings(self) -> Any:
        try:
            import xlwings as xw
        except ImportError as exc:
            raise ExcelServiceError(
                "xlwings is required to run the Excel MCP server."
            ) from exc
        return xw

    def _get_or_create_app(self, *, visible: bool) -> Any:
        app = self._apps.get(visible)
        if app is not None:
            return app

        xw = self._require_xlwings()
        app = xw.App(visible=visible, add_book=False)
        self._apps[visible] = app
        return app

    def _new_workbook_id(self) -> str:
        workbook_id = f"wb_{self._next_workbook_number:03d}"
        self._next_workbook_number += 1
        return workbook_id

    def _session_payload(self, session: WorkbookSession) -> dict[str, JsonValue]:
        try:
            active_sheet = session.workbook.app.selection.sheet.name
        except Exception:
            active_sheet = session.workbook.sheets[0].name
            
        return {
            "workbook_id": session.workbook_id,
            "path": session.path,
            "sheet_names": [sheet.name for sheet in session.workbook.sheets],
            "active_sheet": active_sheet,
            "read_only": session.read_only,
        }

    def _find_open_workbook(self, *, app: Any, target_path: str) -> Any | None:
        for workbook in app.books:
            try:
                workbook_path = str(Path(workbook.fullname).resolve(strict=False))
            except Exception:
                continue
            if workbook_path == target_path:
                return workbook
        return None

    def _get_workbook_session(self, workbook_id: str) -> WorkbookSession:
        session = self._workbooks.get(workbook_id)
        if session is None:
            raise ExcelServiceError(f"Unknown workbook id: {workbook_id}")
        return session

    def _get_sheet(self, *, workbook_id: str, sheet_name: str) -> Any:
        workbook = self._get_workbook_session(workbook_id).workbook
        try:
            return workbook.sheets[sheet_name]
        except Exception as exc:
            raise ExcelServiceError(
                f"Sheet `{sheet_name}` was not found in workbook `{workbook_id}`."
            ) from exc

    def _get_range(self, *, workbook_id: str, sheet_name: str, range_address: str) -> Any:
        worksheet = self._get_sheet(workbook_id=workbook_id, sheet_name=sheet_name)
        try:
            return worksheet.range(range_address)
        except Exception as exc:
            raise ExcelServiceError(
                f"Range `{range_address}` is invalid on sheet `{sheet_name}`."
            ) from exc

    def _get_number_format_matrix(
        self,
        *,
        target_range: Any,
        rows: int,
        cols: int,
    ) -> list[list[str | None]]:
        """Read number formats with a bulk-first strategy and a safe fallback.

        Parameters:
            target_range: The xlwings range to inspect.
            rows: The number of rows in the range.
            cols: The number of columns in the range.

        Returns:
            A 2D matrix of Excel number format strings aligned to the range.
        """

        raw_number_format = read_number_format(target_range)
        if rows == 1 and cols == 1:
            return normalize_number_format_grid(raw_number_format, rows, cols)

        if raw_number_format is not None:
            try:
                return normalize_number_format_grid(raw_number_format, rows, cols)
            except ExcelServiceError:
                pass

        number_format_matrix: list[list[str | None]] = []
        for row_offset in range(rows):
            format_row: list[str | None] = []
            for col_offset in range(cols):
                cell = target_range[row_offset, col_offset]
                cell_number_format = read_number_format(cell)
                format_row.append(normalize_number_format_value(cell_number_format))
            number_format_matrix.append(format_row)

        return number_format_matrix

    def close_all(self) -> None:
        """Safely close all registered workbooks and quit any managed Excel apps."""
        for workbook_id, session in list(self._workbooks.items()):
            try:
                session.workbook.close()
            except Exception:
                pass
        self._workbooks.clear()
        self._path_index.clear()
        
        for visible, app in list(self._apps.items()):
            try:
                app.quit()
            except Exception:
                pass
        self._apps.clear()

    def _recalc_targets(
        self,
        *,
        workbook_id: str,
        scope: str,
        sheet: str | None,
        range_address: str | None,
    ) -> list[Any]:
        session = self._get_workbook_session(workbook_id)
        if scope == "workbook":
            return [worksheet.used_range for worksheet in session.workbook.sheets]
        if scope == "sheet":
            if not sheet:
                raise ExcelServiceError("`sheet` is required when scope is `sheet`.")
            return [self._get_sheet(workbook_id=workbook_id, sheet_name=sheet).used_range]
        if scope == "range":
            if not sheet:
                raise ExcelServiceError("`sheet` is required when scope is `range`.")
            if not range_address:
                raise ExcelServiceError("`range` is required when scope is `range`.")
            return [self._get_range(workbook_id=workbook_id, sheet_name=sheet, range_address=range_address)]
        raise ExcelServiceError("`scope` must be one of `workbook`, `sheet`, or `range`.")

    def _normalize_search_limit(self, limit: int) -> int:
        if isinstance(limit, bool) or not isinstance(limit, int) or limit < 1:
            raise ExcelServiceError("`limit` must be a positive integer.")
        return limit

    def _prepare_search_query(self, query: str | int | float) -> tuple[str, str | int | float]:
        if isinstance(query, bool):
            raise ExcelServiceError("`query` must be a string, int, or float.")
        if isinstance(query, (int, float)):
            return "number", query
        if not isinstance(query, str):
            raise ExcelServiceError("`query` must be a string, int, or float.")
        if query.startswith("="):
            return "formula", normalize_formula_query(query)
        return "text", query.lower()

    def _search_cell_matches(
        self,
        *,
        kind: str,
        prepared_query: str | int | float,
        value: Any,
        formula: Any,
        match_formulas: bool,
    ) -> bool:
        if kind == "number":
            return self._search_number_matches(value=value, query=prepared_query)

        if value not in (None, "") and kind == "text" and prepared_query in str(value).lower():
            return True
        if not isinstance(formula, str):
            return False
        if kind == "formula":
            return prepared_query in normalize_formula_query(formula)
        if kind == "text" and match_formulas:
            return prepared_query in formula.lower()
        return False

    def _search_number_matches(self, *, value: Any, query: str | int | float) -> bool:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            return False
        return abs(float(value) - float(query)) <= 1e-12

    def _normalize_trace_direction(self, direction: str) -> str:
        """Validate and normalize the requested trace direction.

        Parameters:
            direction: The user-provided trace direction string.

        Returns:
            The normalized lowercase direction name.
        """

        normalized_direction = direction.strip().lower()
        if normalized_direction not in {"precedents", "dependents"}:
            raise ExcelServiceError("`direction` must be `precedents` or `dependents`.")
        return normalized_direction

    def _normalize_trace_depth(self, max_depth: int | None) -> int | None:
        """Validate and normalize the requested trace depth.

        Parameters:
            max_depth: The user-provided traversal depth, or ``None``.

        Returns:
            The normalized traversal depth.
        """

        if max_depth is None:
            return None
        if isinstance(max_depth, bool) or not isinstance(max_depth, int) or max_depth < 1:
            raise ExcelServiceError("`max_depth` must be a positive integer or `None`.")
        return max_depth

    def _build_trace_model(self, session: WorkbookSession) -> tuple[Any, str]:
        """Load a formulas workbook model from a snapshot of the live workbook.

        Parameters:
            session: The active workbook session being traced.

        Returns:
            A tuple of ``(trace_model, workbook_name)``.
        """

        formulas = self._require_formulas()
        snapshot_path = self._create_trace_snapshot(session)
        workbook_name = snapshot_path.name
        try:
            with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
                trace_model = formulas.ExcelModel().loads(str(snapshot_path)).finish()
        except Exception as exc:
            raise ExcelServiceError(f"Trace formula could not load workbook snapshot: {exc}") from exc
        finally:
            if snapshot_path != Path(session.path):
                snapshot_path.unlink(missing_ok=True)
        return trace_model, workbook_name

    def _create_sheet_state_snapshot(self, session: WorkbookSession) -> Path:
        """Return a workbook path suitable for fast openpyxl-based sheet analysis.

        Parameters:
            session: The active workbook session whose current state should be read.

        Returns:
            A workbook path on disk. This is the original file when the session is
            already saved, otherwise a temporary SaveCopyAs snapshot.
        """
        workbook_path = Path(session.path)
        try:
            if workbook_path.exists() and bool(session.workbook.api.Saved):
                return workbook_path
        except Exception:
            if workbook_path.exists():
                return workbook_path

        suffix = workbook_path.suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(prefix="excel-mcp-sheet-state-", suffix=suffix, delete=False) as handle:
            snapshot_path = Path(handle.name)

        try:
            session.workbook.api.SaveCopyAs(str(snapshot_path))
            return snapshot_path
        except Exception:
            snapshot_path.unlink(missing_ok=True)

        if workbook_path.exists():
            return workbook_path
        raise ExcelServiceError("Get sheet state could not create a workbook snapshot for analysis.")

    def _create_trace_snapshot(self, session: WorkbookSession) -> Path:
        """Create a temporary workbook snapshot for formulas-based tracing.

        Parameters:
            session: The active workbook session being traced.

        Returns:
            The path to a workbook snapshot on disk.
        """

        suffix = Path(session.path).suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(prefix="excel-mcp-trace-", suffix=suffix, delete=False) as handle:
            snapshot_path = Path(handle.name)

        try:
            session.workbook.api.SaveCopyAs(str(snapshot_path))
            return snapshot_path
        except Exception:
            snapshot_path.unlink(missing_ok=True)

        workbook_path = Path(session.path)
        if workbook_path.exists():
            return workbook_path
        raise ExcelServiceError("Trace formula could not create a workbook snapshot for analysis.")

    def _require_formulas(self) -> Any:
        """Import the formulas package on demand for trace operations.

        Parameters:
            None.

        Returns:
            The imported formulas module.
        """

        try:
            import formulas
        except ImportError as exc:
            raise ExcelServiceError(
                "The `formulas` package is required to trace formula dependencies."
            ) from exc
        return formulas

    def _build_trace_root_refs(
        self,
        *,
        workbook_name: str,
        sheet: str,
        target_range: Any,
    ) -> list[str]:
        """Convert a traced xlwings range into formulas cell refs.

        Parameters:
            workbook_name: The snapshot workbook filename used by formulas.
            sheet: The target sheet name supplied by the caller.
            target_range: The xlwings range requested by the caller.

        Returns:
            A sorted list of workbook-qualified formulas refs for each target cell.
        """

        root_refs: list[str] = []
        sheet_name = sheet.upper()
        start_row = int(target_range.row)
        start_col = int(target_range.column)
        row_count = int(target_range.rows.count)
        col_count = int(target_range.columns.count)
        for row_offset in range(row_count):
            for col_offset in range(col_count):
                root_refs.append(
                    format_formulas_ref(
                        workbook_name,
                        sheet_name,
                        row_column_to_a1_address(start_row + row_offset, start_col + col_offset),
                    )
                )
        return sorted(root_refs)

    def _collect_trace_graph(
        self,
        *,
        trace_model: Any,
        root_refs: list[str],
        direction: str,
        max_depth: int | None,
        active_sheet: str,
        sheet_name_map: dict[str, str],
        include_addresses: bool,
    ) -> tuple[list[dict[str, JsonValue]], list[dict[str, JsonValue]]]:
        """Traverse the formulas model and build a normalized trace graph response.

        Parameters:
            trace_model: The formulas workbook model.
            root_refs: The formulas refs at the traced starting range.
            direction: The normalized trace direction.
            max_depth: The traversal depth cap, or ``None`` for full traversal.
            active_sheet: The user-requested sheet name for display normalization.
            sheet_name_map: Mapping of uppercase sheet names to display names.
            include_addresses: Whether to include split address metadata on nodes.

        Returns:
            A tuple of ``(nodes, edges)`` ready for JSON serialization.
        """

        precedents_by_output: dict[str, set[str]] = {}
        dependents_by_input: dict[str, set[str]] = defaultdict(set)

        for ref, cell in trace_model.cells.items():
            raw_inputs = getattr(cell, 'inputs', None)
            if not raw_inputs:
                continue

            output_ref = str(ref)
            input_refs = {str(input_ref) for input_ref in raw_inputs.keys()}
            precedents_by_output[output_ref] = input_refs
            for input_ref in input_refs:
                dependents_by_input[input_ref].add(output_ref)
                try:
                    expanded_refs = expand_formulas_ref(input_ref)
                except ExcelServiceError:
                    expanded_refs = []
                for expanded_ref in expanded_refs:
                    dependents_by_input[expanded_ref].add(output_ref)

        visited_refs = set(root_refs)
        pending_refs: deque[tuple[str, int]] = deque((ref, 0) for ref in root_refs)
        edge_pairs: set[tuple[str, str]] = set()

        while pending_refs:
            current_ref, depth = pending_refs.popleft()
            if max_depth is not None and depth >= max_depth:
                continue

            if direction == 'precedents':
                next_refs = sorted(precedents_by_output.get(current_ref, set()))
                new_edges = {(next_ref, current_ref) for next_ref in next_refs}
            else:
                next_refs = sorted(dependents_by_input.get(current_ref, set()))
                new_edges = {(current_ref, next_ref) for next_ref in next_refs}

            edge_pairs.update(new_edges)
            for next_ref in next_refs:
                if next_ref not in visited_refs:
                    visited_refs.add(next_ref)
                    pending_refs.append((next_ref, depth + 1))

        node_payloads = [
            build_trace_node_payload(ref, active_sheet, sheet_name_map, include_addresses)
            for ref in sorted(visited_refs, key=lambda ref: normalize_trace_ref(ref, active_sheet, sheet_name_map))
        ]
        normalized_edges = sorted(
            {
                (
                    normalize_trace_ref(source_ref, active_sheet, sheet_name_map),
                    normalize_trace_ref(target_ref, active_sheet, sheet_name_map),
                )
                for source_ref, target_ref in edge_pairs
            }
        )
        edge_payloads = [
            {'from': source_ref, 'to': target_ref}
            for source_ref, target_ref in normalized_edges
        ]
        return node_payloads, edge_payloads


excel_service = ExcelService()
